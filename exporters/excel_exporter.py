"""
exporters/excel_exporter.py

Aggregates all JSON report files in the reports/ directory
and exports them as a single formatted Excel workbook: Summary.xlsx

Design decisions:
    - Reads from reports/ directory → works with any number of machines
    - Each JSON file = one row in the spreadsheet
    - Color-coded cells (green/yellow/red) for instant visual triage
    - Frozen header row and auto-sized columns for usability
    - Overwrites Summary.xlsx each time → it's a derived artifact, not raw data

Why openpyxl?
    - Actively maintained (last release 2024)
    - Full .xlsx support (Excel 2010+)
    - Rich formatting API (colors, borders, fonts, freeze panes)
    - Does NOT require Microsoft Office to be installed
"""

import json
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import REPORTS_DIR, BASE_DIR

logger = logging.getLogger(__name__)

# ── Color Palette ────────────────────────────────────────────────────────────
# Uses ARGB hex strings (Alpha + RGB). Alpha "FF" = fully opaque.

COLORS = {
    "PASS":         "FF2ECC71",   # Emerald green
    "WARNING":      "FFF39C12",   # Orange
    "FAIL":         "FFE74C3C",   # Red
    "UNKNOWN":      "FF95A5A6",   # Grey
    "header_bg":    "FF2C3E50",   # Dark navy
    "header_font":  "FFFFFFFF",   # White
    "row_even":     "FFF8F9FA",   # Light grey
    "row_odd":      "FFFFFFFF",   # White
    "compliant":    "FF27AE60",   # Dark green
    "noncompliant": "FFC0392B",   # Dark red
}

# ── Column Definitions ───────────────────────────────────────────────────────
# Each tuple: (header label, JSON path as list of keys, column width)

COLUMNS: list[tuple[str, list[str], int]] = [
    ("Computer Name",    ["system", "computer_name"],                               22),
    ("Username",         ["system", "username"],                                    18),
    ("Windows Edition",  ["system", "windows_edition"],                             28),
    ("Build Number",     ["system", "build_number"],                                14),
    ("IP Address",       ["system", "ip_address"],                                  16),
    ("MAC Address",      ["system", "mac_address"],                                 20),
    ("RAM (GB)",         ["system", "ram_gb"],                                      12),
    ("Disk (GB)",        ["system", "disk_gb"],                                     12),
    ("Firewall",         ["security", "firewall", "status"],                        12),
    ("Defender",         ["security", "defender", "status"],                        12),
    ("BitLocker",        ["security", "bitlocker", "status"],                       12),
    ("Windows Update",   ["security", "windows_update", "status"],                  16),
    ("WinDefend Svc",    ["services", "WinDefend", "status"],                       16),
    ("BITS Svc",         ["services", "BITS", "status"],                            12),
    ("Windows Upd Svc",  ["services", "wuauserv", "status"],                        16),
    ("Remote Registry",  ["services", "RemoteRegistry", "status"],                  16),
    ("W32Time Svc",      ["services", "W32Time", "status"],                         14),
    ("UAC",              ["registry", "UAC", "status"],                             10),
    ("RDP",              ["registry", "RDP", "status"],                             10),
    ("SMBv1",            ["registry", "SMBv1", "status"],                           12),
    ("Compliance Score", ["compliance_score"],                                       18),
    ("Verdict",          ["compliance", "verdict"],                                  16),
    ("Scan Date",        ["system", "scan_timestamp"],                              22),
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _get_nested(data: dict, keys: list[str], default: str = "N/A") -> str:
    """
    Safely traverse a nested dict using a list of keys.

    Example:
        _get_nested(report, ["security", "firewall", "status"])
        → report["security"]["firewall"]["status"]

    Returns default if any key is missing or value is None.
    """
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current is None:
            return default
    return str(current)


def _status_fill(status: str) -> PatternFill:
    """Return a PatternFill based on PASS/WARNING/FAIL status."""
    color = COLORS.get(status.upper(), COLORS["UNKNOWN"])
    return PatternFill(fill_type="solid", fgColor=color)


def _header_style() -> tuple[Font, PatternFill, Alignment]:
    """Return styles for the header row."""
    font = Font(name="Calibri", bold=True, color=COLORS["header_font"], size=11)
    fill = PatternFill(fill_type="solid", fgColor=COLORS["header_bg"])
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _thin_border() -> Border:
    """Return a thin border for all cells."""
    thin = Side(style="thin", color="FFD5D8DC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _load_reports() -> list[dict]:
    """
    Load all JSON report files from the reports/ directory.

    Skips files that cannot be parsed (corrupted/incomplete).
    Returns a list of report dicts sorted by scan timestamp (oldest first).
    """
    report_files = sorted(REPORTS_DIR.glob("*.json"))

    if not report_files:
        logger.warning("No JSON reports found in: %s", REPORTS_DIR)
        return []

    reports: list[dict] = []
    for filepath in report_files:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
                reports.append(data)
        except (json.JSONDecodeError, OSError) as exc:
            logger.warning("Skipping corrupted report '%s': %s", filepath.name, exc)

    logger.info("Loaded %d report(s) from %s", len(reports), REPORTS_DIR)
    return reports


# ── Main Export Function ─────────────────────────────────────────────────────

def export_summary() -> Path:
    """
    Read all JSON reports and write a formatted Summary.xlsx workbook.

    Workbook structure:
        Sheet "Summary" → one row per scanned machine, color-coded by status.

    Returns:
        Path to the created/updated Summary.xlsx file.

    Raises:
        RuntimeError: If no reports are found to export.
        OSError:      If the file cannot be written.
    """
    reports = _load_reports()

    if not reports:
        raise RuntimeError(
            f"No JSON reports found in '{REPORTS_DIR}'. "
            "Run the auditor on at least one machine first."
        )

    # ── Create workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Summary"

    header_font, header_fill, header_align = _header_style()
    border = _thin_border()

    # ── Header row ───────────────────────────────────────────────────────────
    headers = [col[0] for col in COLUMNS]
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 36

    # ── Data rows ────────────────────────────────────────────────────────────
    STATUS_COLUMNS = {
        "Firewall", "Defender", "BitLocker", "Windows Update",
        "WinDefend Svc", "BITS Svc", "Windows Upd Svc",
        "Remote Registry", "W32Time Svc",
        "UAC", "RDP", "SMBv1",
    }
    VERDICT_COLUMN = "Verdict"

    for row_idx, report in enumerate(reports, start=2):
        # Alternating row background
        row_bg_color = COLORS["row_even"] if row_idx % 2 == 0 else COLORS["row_odd"]
        row_fill = PatternFill(fill_type="solid", fgColor=row_bg_color)

        for col_idx, (header_text, json_path, _) in enumerate(COLUMNS, start=1):
            value = _get_nested(report, json_path)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=10)

            # Color status cells
            if header_text in STATUS_COLUMNS:
                status = value.upper()
                if status in ("PASS", "WARNING", "FAIL", "UNKNOWN"):
                    cell.fill = _status_fill(status)
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
                else:
                    cell.fill = row_fill

            # Color verdict cell
            elif header_text == VERDICT_COLUMN:
                if value == "Compliant":
                    cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["compliant"])
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["noncompliant"])
                    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")

            # Compliance score — bold
            elif header_text == "Compliance Score":
                cell.fill = row_fill
                cell.font = Font(name="Calibri", bold=True, size=10)

            else:
                cell.fill = row_fill

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header row ────────────────────────────────────────────────────
    # Freeze pane at A2 → row 1 stays visible when scrolling down
    ws.freeze_panes = "A2"

    # ── Auto-filter ──────────────────────────────────────────────────────────
    # Allows filtering by any column directly in Excel
    ws.auto_filter.ref = ws.dimensions

    # ── Save ─────────────────────────────────────────────────────────────────
    output_path = BASE_DIR / "Summary.xlsx"
    wb.save(output_path)
    logger.info("Summary.xlsx saved: %s (%d machine(s))", output_path, len(reports))

    return output_path
